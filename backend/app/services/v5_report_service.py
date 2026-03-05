from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

from fastapi import HTTPException, status
from sqlalchemy import Select, func, or_, select
from sqlalchemy.orm import Session, aliased

from app.api.deps import BRIDGE_ADMIN_USERNAME
from app.models.master_data import OilProduct, Warehouse
from app.models.user import User, UserRole
from app.models.v5_domain import (
    Company,
    ContractStatus,
    InventoryMovement,
    PurchaseContract,
    PurchaseContractItem,
    PurchaseOrderV5,
    PurchaseStockIn,
    ReportExport,
    ReportExportStatus,
    ReportType,
    SalesContract,
    SalesContractItem,
    SalesOrderV5,
)
from app.schemas.v5_report import ReportExportItemOut
from app.services.file_storage_service import (
    build_protected_file_url_by_key,
    build_storage_key,
    save_binary_file,
)
from app.services.v5_file_asset_service import ensure_file_asset

_LOCAL_TIMEZONE = ZoneInfo("Asia/Shanghai")
# 频率控制参数：1分钟内最多生成5次，超限后暂停5分钟
_REPORT_RATE_LIMIT_WINDOW_SECONDS = 60
_REPORT_RATE_LIMIT_MAX_COUNT = 5
_REPORT_RATE_LIMIT_PAUSE_SECONDS = 300
_MINIPROGRAM_REPORT_LIST_LIMIT = 5
_REPORT_TITLE_MAP: dict[ReportType, str] = {
    ReportType.SALES_ORDERS: "销售订单报表",
    ReportType.PURCHASE_ORDERS: "采购订单报表",
    ReportType.SALES_CONTRACTS: "销售合同报表",
    ReportType.PURCHASE_CONTRACTS: "采购合同报表",
    ReportType.INVENTORY_MOVEMENTS: "库存变动报表",
    ReportType.WAREHOUSE_LEDGER: "仓库出入库台账",
}
_REPORT_ROLE_MAP: dict[ReportType, set[UserRole]] = {
    ReportType.SALES_ORDERS: {UserRole.CUSTOMER, UserRole.OPERATOR, UserRole.FINANCE, UserRole.ADMIN},
    ReportType.PURCHASE_ORDERS: {UserRole.SUPPLIER, UserRole.OPERATOR, UserRole.FINANCE, UserRole.ADMIN},
    ReportType.SALES_CONTRACTS: {UserRole.CUSTOMER, UserRole.OPERATOR, UserRole.FINANCE, UserRole.ADMIN},
    ReportType.PURCHASE_CONTRACTS: {UserRole.SUPPLIER, UserRole.OPERATOR, UserRole.FINANCE, UserRole.ADMIN},
    ReportType.INVENTORY_MOVEMENTS: {UserRole.OPERATOR, UserRole.FINANCE, UserRole.ADMIN},
    ReportType.WAREHOUSE_LEDGER: {UserRole.WAREHOUSE, UserRole.OPERATOR, UserRole.FINANCE, UserRole.ADMIN},
}


@dataclass(frozen=True)
class ReportDataset:
    headers: list[str]
    rows: list[list[Any]]
    filters: dict[str, Any]
    field_profile: dict[str, Any]
    summary: dict[str, Any]


class ReportGenerateTooFrequentError(ValueError):
    pass


def generate_report_export(
    db: Session,
    *,
    report_type: ReportType,
    current_user: User,
    days: int | None,
    from_date: date | None,
    to_date: date | None,
) -> ReportExport:
    _assert_report_role(report_type=report_type, role=current_user.role)
    _assert_generate_cooldown(db=db, user_id=current_user.id, as_of=datetime.now(UTC))
    resolved_from_date, resolved_to_date = _resolve_date_range(days=days, from_date=from_date, to_date=to_date)
    dataset = _build_dataset(
        db=db,
        report_type=report_type,
        current_user=current_user,
        from_date=resolved_from_date,
        to_date=resolved_to_date,
    )

    workbook_binary = _build_report_workbook(
        title=_REPORT_TITLE_MAP[report_type],
        headers=dataset.headers,
        rows=dataset.rows,
    )
    created_at = datetime.now(UTC)
    created_at_local = created_at.astimezone(_LOCAL_TIMEZONE)
    report_name = f"{_REPORT_TITLE_MAP[report_type]}_{created_at_local.strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_key = build_storage_key(category="report-export", suffix=".xlsx")
    save_binary_file(
        content=workbook_binary,
        file_key=file_key,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        original_filename=report_name,
        allowed_extensions={".xlsx"},
        invalid_extension_detail="文件类型不支持，仅允许 xlsx",
    )
    file_asset = ensure_file_asset(
        db=db,
        file_key=file_key,
        business_type="report_export",
        current_user=current_user,
        file_name=report_name,
    )

    row = ReportExport(
        report_type=report_type,
        generated_by=current_user.id,
        generator_role=current_user.role.value,
        scope_company_id=_resolve_scope_company_id(current_user),
        warehouse_id=None,
        from_date=resolved_from_date,
        to_date=resolved_to_date,
        filters_json=dataset.filters,
        field_profile_json=dataset.field_profile,
        file_asset_id=file_asset.id,
        row_count=len(dataset.rows),
        status=ReportExportStatus.GENERATED,
        summary_json=dataset.summary,
    )
    db.add(row)
    db.flush()
    return row


def list_report_exports(
    db: Session,
    *,
    report_type: ReportType,
    current_user: User,
    days: int | None,
    from_date: date | None,
    to_date: date | None,
    limit: int,
    offset: int,
) -> tuple[list[ReportExport], int]:
    _assert_report_role(report_type=report_type, role=current_user.role)
    resolved_from_date, resolved_to_date = _resolve_date_range(days=days, from_date=from_date, to_date=to_date)

    base_query: Select[tuple[ReportExport]] = select(ReportExport).where(ReportExport.report_type == report_type)
    count_query = select(func.count()).select_from(ReportExport).where(ReportExport.report_type == report_type)
    normalized_limit = limit if _has_global_report_access(current_user) else min(limit, _MINIPROGRAM_REPORT_LIST_LIMIT)

    if not _has_global_report_access(current_user):
        base_query = base_query.where(ReportExport.generated_by == current_user.id)
        count_query = count_query.where(ReportExport.generated_by == current_user.id)

    scope_company_id = _resolve_scope_company_id(current_user)
    if scope_company_id is not None:
        base_query = base_query.where(ReportExport.scope_company_id == scope_company_id)
        count_query = count_query.where(ReportExport.scope_company_id == scope_company_id)

    if resolved_from_date is not None:
        base_query = base_query.where(or_(ReportExport.to_date.is_(None), ReportExport.to_date >= resolved_from_date))
        count_query = count_query.where(or_(ReportExport.to_date.is_(None), ReportExport.to_date >= resolved_from_date))
    if resolved_to_date is not None:
        base_query = base_query.where(or_(ReportExport.from_date.is_(None), ReportExport.from_date <= resolved_to_date))
        count_query = count_query.where(or_(ReportExport.from_date.is_(None), ReportExport.from_date <= resolved_to_date))

    rows = db.scalars(base_query.order_by(ReportExport.created_at.desc()).offset(offset).limit(normalized_limit)).all()
    total = int(db.scalar(count_query) or 0)
    return rows, total


def serialize_report_export_list(db: Session, rows: list[ReportExport]) -> list[ReportExportItemOut]:
    if not rows:
        return []
    user_ids = {item.generated_by for item in rows}
    user_name_map = {
        item.id: (item.display_name or item.username or f"用户{item.id}")
        for item in db.scalars(select(User).where(User.id.in_(user_ids))).all()
    }
    file_asset_ids = {item.file_asset_id for item in rows if item.file_asset_id is not None}
    file_key_map: dict[int, str] = {}
    if file_asset_ids:
        from app.models.v5_domain import FileAsset

        file_key_map = dict(db.execute(select(FileAsset.id, FileAsset.file_key).where(FileAsset.id.in_(file_asset_ids))).all())
    return [
        ReportExportItemOut(
            id=item.id,
            report_type=item.report_type.value,
            report_name=_build_report_name(item.report_type, item.created_at),
            status=item.status.value,
            created_at=item.created_at,
            generated_by=item.generated_by,
            generated_by_name=user_name_map.get(item.generated_by, f"用户{item.generated_by}"),
            row_count=item.row_count,
            download_url=build_protected_file_url_by_key(file_key_map[item.file_asset_id]) if item.file_asset_id in file_key_map else "",
            filters=item.filters_json,
            field_profile=item.field_profile_json,
            summary=item.summary_json,
        )
        for item in rows
    ]


def can_access_report_export_file(db: Session, *, current_user: User, file_key: str) -> bool:
    from app.models.v5_domain import FileAsset

    asset = db.scalar(select(FileAsset).where(FileAsset.file_key == file_key, FileAsset.business_type == "report_export"))
    if asset is None:
        return False
    row = db.scalar(select(ReportExport).where(ReportExport.file_asset_id == asset.id))
    if row is None:
        return False
    if current_user.role not in _REPORT_ROLE_MAP.get(row.report_type, set()):
        return False
    if _has_global_report_access(current_user):
        return True
    if row.generated_by != current_user.id:
        return False
    scope_company_id = _resolve_scope_company_id(current_user, allow_missing=True)
    if scope_company_id is not None:
        return row.scope_company_id == scope_company_id
    return True


def _has_global_report_access(current_user: User) -> bool:
    return current_user.username == BRIDGE_ADMIN_USERNAME


def _assert_report_role(*, report_type: ReportType, role: UserRole) -> None:
    allowed_roles = _REPORT_ROLE_MAP[report_type]
    if role not in allowed_roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="forbidden")


def _assert_generate_cooldown(db: Session, *, user_id: int, as_of: datetime) -> None:
    """
    频率控制：1分钟内最多生成5次，超限后触发5分钟暂停期。
    算法：查询最近6分钟（1分钟窗口 + 5分钟暂停）内的生成记录，
    用滑动窗口检测是否曾出现"1分钟内5次"的情况，若找到则计算暂停结束时间。
    若仍在暂停期则拦截；否则检查当前1分钟内已有计数是否达到上限。
    """
    six_minutes_ago = as_of - timedelta(seconds=_REPORT_RATE_LIMIT_WINDOW_SECONDS + _REPORT_RATE_LIMIT_PAUSE_SECONDS)
    rows = db.scalars(
        select(ReportExport)
        .where(ReportExport.generated_by == user_id)
        .where(ReportExport.created_at >= six_minutes_ago)
        .order_by(ReportExport.created_at.desc())
    ).all()
    if not rows:
        return

    def _aware(t: datetime) -> datetime:
        return t if t.tzinfo is not None else t.replace(tzinfo=UTC)

    # 检测6分钟窗口内是否存在"连续5条记录跨度≤1分钟"的突发情况，找最近一次
    pause_until: datetime | None = None
    for i in range(len(rows) - (_REPORT_RATE_LIMIT_MAX_COUNT - 1)):
        newest_t = _aware(rows[i].created_at)
        fifth_t = _aware(rows[i + _REPORT_RATE_LIMIT_MAX_COUNT - 1].created_at)
        if newest_t - fifth_t <= timedelta(seconds=_REPORT_RATE_LIMIT_WINDOW_SECONDS):
            # 发现1分钟内5次的突发，暂停截止时间 = 最新那条 + 5分钟
            candidate = newest_t + timedelta(seconds=_REPORT_RATE_LIMIT_PAUSE_SECONDS)
            if pause_until is None or candidate > pause_until:
                pause_until = candidate
            break  # rows 按倒序排列，第一个匹配即为最新突发

    if pause_until is not None and as_of < pause_until:
        raise ReportGenerateTooFrequentError("生成过于频繁，请5分钟后再试")

    # 暂停期已过或从未超限：检查当前1分钟内已生成次数是否达到上限
    one_minute_ago = as_of - timedelta(seconds=_REPORT_RATE_LIMIT_WINDOW_SECONDS)
    count_in_minute = sum(1 for r in rows if _aware(r.created_at) >= one_minute_ago)
    if count_in_minute >= _REPORT_RATE_LIMIT_MAX_COUNT:
        raise ReportGenerateTooFrequentError("生成过于频繁，请稍后再试")


def _resolve_scope_company_id(current_user: User, *, allow_missing: bool = False) -> int | None:
    if current_user.role in {UserRole.CUSTOMER, UserRole.SUPPLIER, UserRole.WAREHOUSE}:
        if current_user.company_id is None and not allow_missing:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="company_scope_missing")
        return current_user.company_id
    return None


def _resolve_date_range(*, days: int | None, from_date: date | None, to_date: date | None) -> tuple[date | None, date | None]:
    if from_date and to_date and from_date > to_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_date_range")
    if from_date is not None or to_date is not None:
        return from_date, to_date
    if days is None or days == 0:
        return None, None
    if days < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid_date_range")
    today = datetime.now(UTC).astimezone(_LOCAL_TIMEZONE).date()
    return today - timedelta(days=days - 1), today


def _build_dataset(
    db: Session,
    *,
    report_type: ReportType,
    current_user: User,
    from_date: date | None,
    to_date: date | None,
) -> ReportDataset:
    if report_type == ReportType.SALES_ORDERS:
        return _build_sales_order_dataset(db=db, current_user=current_user, from_date=from_date, to_date=to_date)
    if report_type == ReportType.PURCHASE_ORDERS:
        return _build_purchase_order_dataset(db=db, current_user=current_user, from_date=from_date, to_date=to_date)
    if report_type == ReportType.SALES_CONTRACTS:
        return _build_sales_contract_dataset(db=db, current_user=current_user, from_date=from_date, to_date=to_date)
    if report_type == ReportType.PURCHASE_CONTRACTS:
        return _build_purchase_contract_dataset(db=db, current_user=current_user, from_date=from_date, to_date=to_date)
    if report_type == ReportType.INVENTORY_MOVEMENTS:
        return _build_inventory_movement_dataset(db=db, from_date=from_date, to_date=to_date)
    if report_type == ReportType.WAREHOUSE_LEDGER:
        return _build_warehouse_ledger_dataset(db=db, current_user=current_user, from_date=from_date, to_date=to_date)
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="unsupported_report_type")


def _build_sales_order_dataset(
    db: Session,
    *,
    current_user: User,
    from_date: date | None,
    to_date: date | None,
) -> ReportDataset:
    customer_company = aliased(Company)
    supplier_company = aliased(Company)
    warehouse = aliased(Warehouse)
    product = aliased(OilProduct)
    sales_contract = aliased(SalesContract)
    purchase_contract = aliased(PurchaseContract)
    purchase_order = aliased(PurchaseOrderV5)

    query = (
        select(
            SalesOrderV5.order_date,
            SalesOrderV5.sales_order_no,
            customer_company.company_name.label("customer_company_name"),
            SalesOrderV5.operator_company_name_snapshot,
            warehouse.name.label("warehouse_name"),
            product.name.label("product_name"),
            SalesOrderV5.qty_ton,
            sales_contract.contract_no.label("sales_contract_no"),
            SalesOrderV5.amount_tax_included.label("sales_amount"),
            SalesOrderV5.received_amount,
            SalesOrderV5.status,
            purchase_order.purchase_order_no,
            supplier_company.company_name.label("supplier_company_name"),
            purchase_contract.contract_no.label("purchase_contract_no"),
            purchase_order.amount_tax_included.label("purchase_amount"),
        )
        .join(customer_company, customer_company.id == SalesOrderV5.customer_company_id)
        .join(warehouse, warehouse.id == SalesOrderV5.warehouse_id)
        .join(product, product.id == SalesOrderV5.product_id)
        .join(sales_contract, sales_contract.id == SalesOrderV5.sales_contract_id)
        .outerjoin(purchase_order, purchase_order.sales_order_id == SalesOrderV5.id)
        .outerjoin(purchase_contract, purchase_contract.id == purchase_order.purchase_contract_id)
        .outerjoin(supplier_company, supplier_company.id == purchase_order.supplier_company_id)
    )
    if current_user.role == UserRole.CUSTOMER:
        query = query.where(SalesOrderV5.customer_company_id == current_user.company_id)
    if from_date is not None:
        query = query.where(SalesOrderV5.order_date >= from_date)
    if to_date is not None:
        query = query.where(SalesOrderV5.order_date <= to_date)
    rows = db.execute(query.order_by(SalesOrderV5.order_date.desc(), SalesOrderV5.id.desc())).all()

    common_headers = [
        "订单日期",
        "销售订单号",
        "客户公司",
        "运营公司",
        "仓库",
        "油品",
        "数量(吨)",
        "销售合同号",
        "销售金额",
        "收款金额",
        "状态",
        "采购订单号",
    ]
    common_field_ids = [
        "order_date",
        "sales_order_no",
        "customer_company_name",
        "operator_company_name",
        "warehouse_name",
        "product_name",
        "qty_ton",
        "sales_contract_no",
        "sales_amount",
        "received_amount",
        "status",
        "purchase_order_no",
    ]
    extra_headers = ["供应商公司", "采购合同号", "采购金额"] if current_user.role != UserRole.CUSTOMER else []
    extra_field_ids = ["supplier_company_name", "purchase_contract_no", "purchase_amount"] if current_user.role != UserRole.CUSTOMER else []

    data_rows = []
    for item in rows:
        current_row = [
            item.order_date.isoformat(),
            item.sales_order_no,
            item.customer_company_name or "",
            item.operator_company_name_snapshot or "",
            item.warehouse_name or "",
            item.product_name or "",
            _decimal_to_float(item.qty_ton),
            item.sales_contract_no or "",
            _decimal_to_float(item.sales_amount),
            _decimal_to_float(item.received_amount),
            item.status.value,
            item.purchase_order_no or "",
        ]
        if current_user.role != UserRole.CUSTOMER:
            current_row.extend(
                [
                    item.supplier_company_name or "",
                    item.purchase_contract_no or "",
                    _decimal_to_float(item.purchase_amount),
                ]
            )
        data_rows.append(current_row)

    return ReportDataset(
        headers=common_headers + extra_headers,
        rows=data_rows,
        filters={"from_date": _date_text(from_date), "to_date": _date_text(to_date)},
        field_profile={"field_ids": common_field_ids + extra_field_ids, "role": current_user.role.value},
        summary={"row_count": len(data_rows)},
    )


def _build_purchase_order_dataset(
    db: Session,
    *,
    current_user: User,
    from_date: date | None,
    to_date: date | None,
) -> ReportDataset:
    supplier_company = aliased(Company)
    customer_company = aliased(Company)
    warehouse = aliased(Warehouse)
    product = aliased(OilProduct)
    purchase_contract = aliased(PurchaseContract)
    sales_contract = aliased(SalesContract)
    sales_order = aliased(SalesOrderV5)

    query = (
        select(
            sales_order.order_date,
            PurchaseOrderV5.purchase_order_no,
            supplier_company.company_name.label("supplier_company_name"),
            warehouse.name.label("warehouse_name"),
            product.name.label("product_name"),
            PurchaseOrderV5.qty_ton,
            purchase_contract.contract_no.label("purchase_contract_no"),
            PurchaseOrderV5.amount_tax_included.label("purchase_amount"),
            PurchaseOrderV5.status,
            customer_company.company_name.label("customer_company_name"),
            sales_order.sales_order_no,
            sales_contract.contract_no.label("sales_contract_no"),
            sales_order.amount_tax_included.label("sales_amount"),
        )
        .join(sales_order, sales_order.id == PurchaseOrderV5.sales_order_id)
        .outerjoin(supplier_company, supplier_company.id == PurchaseOrderV5.supplier_company_id)
        .join(warehouse, warehouse.id == PurchaseOrderV5.warehouse_id)
        .join(product, product.id == PurchaseOrderV5.product_id)
        .outerjoin(purchase_contract, purchase_contract.id == PurchaseOrderV5.purchase_contract_id)
        .join(customer_company, customer_company.id == sales_order.customer_company_id)
        .join(sales_contract, sales_contract.id == sales_order.sales_contract_id)
    )
    if current_user.role == UserRole.SUPPLIER:
        query = query.where(PurchaseOrderV5.supplier_company_id == current_user.company_id)
    if from_date is not None:
        query = query.where(sales_order.order_date >= from_date)
    if to_date is not None:
        query = query.where(sales_order.order_date <= to_date)
    rows = db.execute(query.order_by(sales_order.order_date.desc(), PurchaseOrderV5.id.desc())).all()

    common_headers = [
        "订单日期",
        "采购订单号",
        "供应商公司",
        "仓库",
        "油品",
        "数量(吨)",
        "采购合同号",
        "采购金额",
        "状态",
    ]
    common_field_ids = [
        "order_date",
        "purchase_order_no",
        "supplier_company_name",
        "warehouse_name",
        "product_name",
        "qty_ton",
        "purchase_contract_no",
        "purchase_amount",
        "status",
    ]
    extra_headers = ["客户公司", "销售订单号", "销售合同号", "销售金额"] if current_user.role != UserRole.SUPPLIER else []
    extra_field_ids = ["customer_company_name", "sales_order_no", "sales_contract_no", "sales_amount"] if current_user.role != UserRole.SUPPLIER else []

    data_rows = []
    for item in rows:
        current_row = [
            item.order_date.isoformat(),
            item.purchase_order_no,
            item.supplier_company_name or "",
            item.warehouse_name or "",
            item.product_name or "",
            _decimal_to_float(item.qty_ton),
            item.purchase_contract_no or "",
            _decimal_to_float(item.purchase_amount),
            item.status.value,
        ]
        if current_user.role != UserRole.SUPPLIER:
            current_row.extend(
                [
                    item.customer_company_name or "",
                    item.sales_order_no or "",
                    item.sales_contract_no or "",
                    _decimal_to_float(item.sales_amount),
                ]
            )
        data_rows.append(current_row)

    return ReportDataset(
        headers=common_headers + extra_headers,
        rows=data_rows,
        filters={"from_date": _date_text(from_date), "to_date": _date_text(to_date)},
        field_profile={"field_ids": common_field_ids + extra_field_ids, "role": current_user.role.value},
        summary={"row_count": len(data_rows)},
    )


def _build_sales_contract_dataset(
    db: Session,
    *,
    current_user: User,
    from_date: date | None,
    to_date: date | None,
) -> ReportDataset:
    amount_map = _load_contract_amount_map(db=db, contract_type="SALES")
    customer_company = aliased(Company)
    query = (
        select(SalesContract, customer_company.company_name)
        .join(customer_company, customer_company.id == SalesContract.customer_company_id)
    )
    if current_user.role == UserRole.CUSTOMER:
        query = query.where(SalesContract.customer_company_id == current_user.company_id)
    if from_date is not None:
        query = query.where(SalesContract.contract_date >= from_date)
    if to_date is not None:
        query = query.where(SalesContract.contract_date <= to_date)
    rows = db.execute(query.order_by(SalesContract.contract_date.desc(), SalesContract.id.desc())).all()

    headers = [
        "合同日期",
        "销售合同号",
        "客户公司",
        "状态",
        "合同数量",
        "已执行数量",
        "待执行数量",
        "超执行数量",
        "保证金金额",
        "合同金额",
        "生效时间",
    ]
    field_ids = [
        "contract_date",
        "contract_no",
        "customer_company_name",
        "status",
        "effective_contract_qty",
        "executed_qty",
        "pending_execution_qty",
        "over_executed_qty",
        "deposit_amount",
        "amount_tax_included",
        "effective_at",
    ]
    data_rows = [
        [
            contract.contract_date.isoformat(),
            contract.contract_no,
            company_name or "",
            contract.status.value,
            _decimal_to_float(contract.effective_contract_qty),
            _decimal_to_float(contract.executed_qty),
            _decimal_to_float(contract.pending_execution_qty),
            _decimal_to_float(contract.over_executed_qty),
            _decimal_to_float(contract.deposit_amount),
            _decimal_to_float(amount_map.get(contract.id)),
            _datetime_text(contract.effective_at),
        ]
        for contract, company_name in rows
    ]
    return ReportDataset(
        headers=headers,
        rows=data_rows,
        filters={"from_date": _date_text(from_date), "to_date": _date_text(to_date)},
        field_profile={"field_ids": field_ids, "role": current_user.role.value},
        summary={"row_count": len(data_rows)},
    )


def _build_purchase_contract_dataset(
    db: Session,
    *,
    current_user: User,
    from_date: date | None,
    to_date: date | None,
) -> ReportDataset:
    amount_map = _load_contract_amount_map(db=db, contract_type="PURCHASE")
    supplier_company = aliased(Company)
    query = (
        select(PurchaseContract, supplier_company.company_name)
        .join(supplier_company, supplier_company.id == PurchaseContract.supplier_company_id)
    )
    if current_user.role == UserRole.SUPPLIER:
        query = query.where(PurchaseContract.supplier_company_id == current_user.company_id)
    if from_date is not None:
        query = query.where(PurchaseContract.contract_date >= from_date)
    if to_date is not None:
        query = query.where(PurchaseContract.contract_date <= to_date)
    rows = db.execute(query.order_by(PurchaseContract.contract_date.desc(), PurchaseContract.id.desc())).all()

    warehouse_ids = {
        int(contract.template_snapshot_json.get("warehouse_id"))
        for contract, _ in rows
        if isinstance(contract.template_snapshot_json, dict) and contract.template_snapshot_json.get("warehouse_id")
    }
    warehouse_name_map = dict(db.execute(select(Warehouse.id, Warehouse.name).where(Warehouse.id.in_(warehouse_ids))).all()) if warehouse_ids else {}

    headers = [
        "合同日期",
        "采购合同号",
        "供应商公司",
        "仓库",
        "状态",
        "合同数量",
        "已入库数量",
        "待入库数量",
        "已执行数量",
        "待执行数量",
        "超执行数量",
        "保证金金额",
        "合同金额",
        "生效时间",
    ]
    field_ids = [
        "contract_date",
        "contract_no",
        "supplier_company_name",
        "warehouse_name",
        "status",
        "effective_contract_qty",
        "stocked_in_qty",
        "pending_stock_in_qty",
        "executed_qty",
        "pending_execution_qty",
        "over_executed_qty",
        "deposit_amount",
        "amount_tax_included",
        "effective_at",
    ]
    data_rows = []
    for contract, company_name in rows:
        warehouse_id = None
        if isinstance(contract.template_snapshot_json, dict) and contract.template_snapshot_json.get("warehouse_id") is not None:
            warehouse_id = int(contract.template_snapshot_json["warehouse_id"])
        data_rows.append(
            [
                contract.contract_date.isoformat(),
                contract.contract_no,
                company_name or "",
                warehouse_name_map.get(warehouse_id, "") if warehouse_id is not None else "",
                contract.status.value,
                _decimal_to_float(contract.effective_contract_qty),
                _decimal_to_float(contract.stocked_in_qty),
                _decimal_to_float(contract.pending_stock_in_qty),
                _decimal_to_float(contract.executed_qty),
                _decimal_to_float(contract.pending_execution_qty),
                _decimal_to_float(contract.over_executed_qty),
                _decimal_to_float(contract.deposit_amount),
                _decimal_to_float(amount_map.get(contract.id)),
                _datetime_text(contract.effective_at),
            ]
        )
    return ReportDataset(
        headers=headers,
        rows=data_rows,
        filters={"from_date": _date_text(from_date), "to_date": _date_text(to_date)},
        field_profile={"field_ids": field_ids, "role": current_user.role.value},
        summary={"row_count": len(data_rows)},
    )


def _build_inventory_movement_dataset(
    db: Session,
    *,
    from_date: date | None,
    to_date: date | None,
) -> ReportDataset:
    query = select(InventoryMovement).order_by(InventoryMovement.created_at.desc(), InventoryMovement.id.desc())
    query = _apply_datetime_date_range(query, field=InventoryMovement.created_at, from_date=from_date, to_date=to_date)
    rows = db.scalars(query).all()
    warehouse_name_map, product_name_map, user_name_map = _build_reference_maps_for_movements(db=db, rows=rows)

    headers = [
        "流水时间",
        "流水号",
        "仓库",
        "油品",
        "变动类型",
        "业务类型",
        "业务单据ID",
        "变动前库存",
        "变动数量",
        "变动后库存",
        "变动前冻结",
        "变动后冻结",
        "操作人",
        "备注",
    ]
    field_ids = [
        "created_at",
        "movement_no",
        "warehouse_name",
        "product_name",
        "movement_type",
        "business_type",
        "business_id",
        "before_on_hand_qty_ton",
        "change_qty_ton",
        "after_on_hand_qty_ton",
        "before_reserved_qty_ton",
        "after_reserved_qty_ton",
        "operator_name",
        "remark",
    ]
    data_rows = [
        [
            _datetime_text(item.created_at),
            item.movement_no,
            warehouse_name_map.get(item.warehouse_id, ""),
            product_name_map.get(item.product_id, ""),
            item.movement_type.value,
            item.business_type,
            item.business_id,
            _decimal_to_float(item.before_on_hand_qty_ton),
            _decimal_to_float(item.change_qty_ton),
            _decimal_to_float(item.after_on_hand_qty_ton),
            _decimal_to_float(item.before_reserved_qty_ton),
            _decimal_to_float(item.after_reserved_qty_ton),
            user_name_map.get(item.operator_user_id, f"用户{item.operator_user_id}"),
            item.remark or "",
        ]
        for item in rows
    ]
    return ReportDataset(
        headers=headers,
        rows=data_rows,
        filters={"from_date": _date_text(from_date), "to_date": _date_text(to_date)},
        field_profile={"field_ids": field_ids},
        summary={"row_count": len(data_rows)},
    )


def _build_warehouse_ledger_dataset(
    db: Session,
    *,
    current_user: User,
    from_date: date | None,
    to_date: date | None,
) -> ReportDataset:
    query = select(InventoryMovement).order_by(InventoryMovement.created_at.desc(), InventoryMovement.id.desc())
    query = _apply_datetime_date_range(query, field=InventoryMovement.created_at, from_date=from_date, to_date=to_date)
    if current_user.role == UserRole.WAREHOUSE:
        warehouse_ids = db.scalars(select(Warehouse.id).where(Warehouse.company_id == current_user.company_id)).all()
        if warehouse_ids:
            query = query.where(InventoryMovement.warehouse_id.in_(warehouse_ids))
        else:
            query = query.where(InventoryMovement.id == -1)
    rows = db.scalars(query).all()
    warehouse_name_map, product_name_map, _ = _build_reference_maps_for_movements(db=db, rows=rows)
    purchase_order_no_map, purchase_contract_no_map = _build_ledger_reference_maps(db=db, rows=rows)

    headers = [
        "流水时间",
        "流水号",
        "仓库",
        "油品",
        "变动类型",
        "数量(吨)",
        "关联采购订单号",
        "关联采购合同号",
        "备注",
    ]
    field_ids = [
        "created_at",
        "movement_no",
        "warehouse_name",
        "product_name",
        "movement_type",
        "change_qty_ton",
        "purchase_order_no",
        "purchase_contract_no",
        "remark",
    ]
    data_rows = [
        [
            _datetime_text(item.created_at),
            item.movement_no,
            warehouse_name_map.get(item.warehouse_id, ""),
            product_name_map.get(item.product_id, ""),
            item.movement_type.value,
            _decimal_to_float(item.change_qty_ton),
            purchase_order_no_map.get(item.id, ""),
            purchase_contract_no_map.get(item.id, ""),
            item.remark or "",
        ]
        for item in rows
    ]
    return ReportDataset(
        headers=headers,
        rows=data_rows,
        filters={"from_date": _date_text(from_date), "to_date": _date_text(to_date)},
        field_profile={"field_ids": field_ids, "role": current_user.role.value},
        summary={"row_count": len(data_rows)},
    )


def _load_contract_amount_map(db: Session, *, contract_type: str) -> dict[int, Decimal]:
    if contract_type == "SALES":
        rows = db.execute(
            select(SalesContractItem.sales_contract_id, func.coalesce(func.sum(SalesContractItem.amount_tax_included), 0))
            .group_by(SalesContractItem.sales_contract_id)
        ).all()
        return {int(contract_id): amount for contract_id, amount in rows}
    rows = db.execute(
        select(PurchaseContractItem.purchase_contract_id, func.coalesce(func.sum(PurchaseContractItem.amount_tax_included), 0))
        .group_by(PurchaseContractItem.purchase_contract_id)
    ).all()
    return {int(contract_id): amount for contract_id, amount in rows}


def _build_reference_maps_for_movements(
    db: Session,
    *,
    rows: list[InventoryMovement],
) -> tuple[dict[int, str], dict[int, str], dict[int, str]]:
    warehouse_ids = {item.warehouse_id for item in rows}
    product_ids = {item.product_id for item in rows}
    user_ids = {item.operator_user_id for item in rows}
    warehouse_name_map = dict(db.execute(select(Warehouse.id, Warehouse.name).where(Warehouse.id.in_(warehouse_ids))).all()) if warehouse_ids else {}
    product_name_map = dict(db.execute(select(OilProduct.id, OilProduct.name).where(OilProduct.id.in_(product_ids))).all()) if product_ids else {}
    user_name_map = {
        item.id: (item.display_name or item.username or f"用户{item.id}")
        for item in db.scalars(select(User).where(User.id.in_(user_ids))).all()
    } if user_ids else {}
    return warehouse_name_map, product_name_map, user_name_map


def _build_ledger_reference_maps(
    db: Session,
    *,
    rows: list[InventoryMovement],
) -> tuple[dict[int, str], dict[int, str]]:
    purchase_order_ids = {item.business_id for item in rows if item.business_type == "PURCHASE_ORDER"}
    stock_in_ids = {item.business_id for item in rows if item.business_type == "PURCHASE_STOCK_IN"}

    purchase_orders = db.scalars(select(PurchaseOrderV5).where(PurchaseOrderV5.id.in_(purchase_order_ids))).all() if purchase_order_ids else []
    purchase_order_map = {item.id: item for item in purchase_orders}
    stock_ins = db.scalars(select(PurchaseStockIn).where(PurchaseStockIn.id.in_(stock_in_ids))).all() if stock_in_ids else []
    stock_in_map = {item.id: item for item in stock_ins}

    purchase_contract_ids = {
        item.purchase_contract_id
        for item in purchase_orders
        if item.purchase_contract_id is not None
    } | {
        item.purchase_contract_id
        for item in stock_ins
    }
    purchase_contract_no_map = dict(
        db.execute(select(PurchaseContract.id, PurchaseContract.contract_no).where(PurchaseContract.id.in_(purchase_contract_ids))).all()
    ) if purchase_contract_ids else {}

    movement_purchase_order_no_map: dict[int, str] = {}
    movement_purchase_contract_no_map: dict[int, str] = {}
    for item in rows:
        if item.business_type == "PURCHASE_ORDER":
            purchase_order = purchase_order_map.get(item.business_id)
            if purchase_order is not None:
                movement_purchase_order_no_map[item.id] = purchase_order.purchase_order_no
                if purchase_order.purchase_contract_id is not None:
                    movement_purchase_contract_no_map[item.id] = purchase_contract_no_map.get(purchase_order.purchase_contract_id, "")
            continue
        if item.business_type == "PURCHASE_STOCK_IN":
            stock_in = stock_in_map.get(item.business_id)
            if stock_in is not None:
                movement_purchase_contract_no_map[item.id] = purchase_contract_no_map.get(stock_in.purchase_contract_id, "")
    return movement_purchase_order_no_map, movement_purchase_contract_no_map


def _apply_datetime_date_range(
    query: Select[tuple[InventoryMovement]],
    *,
    field: Any,
    from_date: date | None,
    to_date: date | None,
) -> Select[tuple[InventoryMovement]]:
    if from_date is not None:
        query = query.where(field >= datetime.combine(from_date, time.min, tzinfo=UTC))
    if to_date is not None:
        query = query.where(field < datetime.combine(to_date + timedelta(days=1), time.min, tzinfo=UTC))
    return query


def _build_report_workbook(*, title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("未安装 openpyxl，无法生成 V5 报表") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title[:31]
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="D9E8FB")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    for index, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=index).value
            max_len = max(max_len, len(str(value or "")))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 12), 28)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_report_name(report_type: ReportType, created_at: datetime) -> str:
    local_dt = created_at
    if created_at.tzinfo is None:
        local_dt = created_at.replace(tzinfo=UTC)
    local_dt = local_dt.astimezone(_LOCAL_TIMEZONE)
    return f"{_REPORT_TITLE_MAP[report_type]}_{local_dt.strftime('%Y%m%d_%H%M%S')}.xlsx"


def _date_text(value: date | None) -> str | None:
    return value.isoformat() if value is not None else None


def _datetime_text(value: datetime | None) -> str:
    if value is None:
        return ""
    current = value
    if current.tzinfo is None:
        current = current.replace(tzinfo=UTC)
    return current.astimezone(_LOCAL_TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")


def _decimal_to_float(value: Decimal | None) -> float:
    if value is None:
        return 0.0
    return float(value)
